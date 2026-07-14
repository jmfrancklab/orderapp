"""ACERT order interface — Flask + SQLite.

Single-file backend (plus quotes.py for Dropbox/SharePoint quote handling).
All state lives in orders.db next to this file (absolute path, so it works
identically under PythonAnywhere's WSGI).
"""
import os
import re
import sqlite3
import tomllib
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse

from flask import (Flask, g, jsonify, redirect, render_template, request,
                   session, url_for)
from werkzeug.middleware.proxy_fix import ProxyFix

import quotes

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "orders.db")

# Increment this (major.minor.patch) whenever you deploy a meaningful change.
__version__ = "0.10.9"

# ── Config ────────────────────────────────────────────────────────────────────
def _load_config():
    path = os.path.join(BASE_DIR, "config.toml")
    if os.path.exists(path):
        with open(path, "rb") as f:
            return tomllib.load(f)
    return {}

_CONFIG = _load_config()

app = Flask(__name__)
# Fix HTTPS scheme detection behind PythonAnywhere's reverse proxy.
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
# CHANGE THIS before deploying (any long random string):
app.secret_key = os.environ.get("ORDERAPP_SECRET", "dev-secret-change-me")


@app.template_filter('fmt_cost')
def fmt_cost(value):
    """Format a stored cost string ('1234.56') as '1,234.56' for display."""
    s = str(value or '').replace(',', '').strip()
    if not s:
        return ''
    try:
        return '{:,.2f}'.format(float(s))
    except ValueError:
        return value or ''


def _normalise_cost(raw):
    """Strip commas/whitespace from user input; return plain decimal or ''."""
    s = str(raw or '').replace(',', '').strip()
    if not s:
        return ''
    try:
        return str(round(float(s), 2))
    except ValueError:
        return s


# ── Bookmarklet ───────────────────────────────────────────────────────────────

# Minified bookmarklet — __CAPTURE_URL__ is substituted with the real endpoint
# URL at render time inside inject_globals().
_BOOKMARKLET = (
    r"(function(){"
    r"var url=window.location.href;"
    r"var c=document.querySelector('link[rel=\"canonical\"]');"
    r"if(c&&c.href)url=c.href;"
    r"else{var o=document.querySelector('meta[property=\"og:url\"]');"
    r"if(o&&o.content)url=o.content;}"
    r"var price=null,desc=null,vname=null,addr=null,phone=null,website=null;"
    r"[].forEach.call(document.querySelectorAll('[type=\"application/ld+json\"]'),function(s){"
    r"try{var d=JSON.parse(s.textContent);"
    r"(Array.isArray(d)?d:[d]).forEach(function(item){"
    r"var t=(item['@type']||'').toString();"
    r"if(/Product/.test(t)){"
    r"if(!desc)desc=item.name||null;"
    r"var o=item.offers;if(Array.isArray(o))o=o.filter(function(x){return x&&x.price;})[0];"
    r"if(o&&o.price&&!price)price=String(o.price);"
    r"if(item.brand&&item.brand.name&&!vname)vname=item.brand.name;}"
    r"if(/Organization|LocalBusiness|Corporation/.test(t)){"
    r"if(!vname&&item.name)vname=item.name;"
    r"if(!phone&&item.telephone)phone=item.telephone;"
    r"if(!website&&item.url)website=item.url;"
    r"var a=item.address;"
    r"if(a&&typeof a==='object'&&!addr){"
    r"addr=[a.streetAddress,[a.addressLocality,a.addressRegion,a.postalCode]"
    r".filter(Boolean).join(', ')].filter(Boolean).join(', ');}}"
    r"});}"
    r"catch(e){}});"
    r"if(!price){"
    r"['.pd-price','[itemprop=\"price\"]','.price','#priceblock_ourprice',"
    r"'.a-price .a-offscreen','span[data-a-color=\"price\"] .a-offscreen',"
    r"'.priceToPay .a-offscreen'].some(function(sel){"
    r"var el=document.querySelector(sel);"
    r"if(el){var m=(el.textContent||el.getAttribute('content')||'').match(/([\d,]+\.\d{2})/);"
    r"if(m){price=m[1].replace(/,/g,'');return true;}}});}"
    r"if(!price){"
    r"var w=document.querySelector('.a-price-whole');"
    r"var f=document.querySelector('.a-price-fraction');"
    r"if(w&&f)price=(w.textContent.replace(/[^0-9]/g,''))+'.'+(f.textContent.replace(/[^0-9]/g,''));}"
    r"if(!price){"
    # Scan raw page source for embedded price JSON (catches eBay /p/ pages and
    # other sites where price is in a non-JSON-LD script block).
    r"var src=document.documentElement.innerHTML;"
    r"var pm=src.match(/['\"]price['\"]\s*:\s*['\"]?([\d]+\.[\d]{2})['\"]?/);"
    r"if(pm&&parseFloat(pm[1])>0)price=pm[1];}"
    r"if(!desc)desc=document.title||'';"
    r"if(!website)website=window.location.hostname;"
    r"var div=document.createElement('div');"
    r"div.style.cssText='position:fixed;top:10px;right:10px;background:#0e6e6b;color:#fff;"
    r"padding:12px 18px;border-radius:4px;font:600 14px system-ui;z-index:99999;"
    r"box-shadow:0 4px 12px rgba(0,0,0,.3);transition:opacity .3s;';"
    r"div.textContent='⏳ Sending to ACERT…';"
    r"document.body.appendChild(div);"
    # Use window.open relay so the POST happens same-origin on our server,
    # bypassing any CSP connect-src restrictions on the product page (e.g. eBay).
    # The relay page postMessages back the result so we can update the overlay.
    r"var payload=btoa(unescape(encodeURIComponent(JSON.stringify("
    r"{url:url,price:price,description:desc,vendor_name:vname,"
    r"address:addr,phone:phone,website:website}))));"
    r"var timer=setTimeout(function(){"
    r"div.textContent=price?'✓ Sent: $'+price:'✓ Sent!';"
    r"setTimeout(function(){div.remove();},1800);},8000);"
    r"window.addEventListener('message',function(e){"
    r"if(!e.data||e.data.type==='acert-ok'&&!e.data.price&&!price)return;"
    r"if(e.data.type==='acert-ok'||e.data.type==='acert-err'){"
    r"clearTimeout(timer);"
    r"if(e.data.type==='acert-ok'){"
    r"div.textContent=price?'✓ Captured: $'+price:'✓ Captured!';"
    r"setTimeout(function(){div.remove();},1800);}"
    r"else{div.style.background='#c0392b';"
    r"div.textContent='✗ '+(e.data.error||'Not captured — sign in to ACERT first');"
    r"setTimeout(function(){div.remove();},4000);}}},"
    r"{once:true});"
    r"window.open('__RELAY_URL__#'+payload);"
    r"})()"
)

# In-memory capture store: email → list of capture dicts from the bookmarklet.
# Cleared when the orders page reads them. Intentionally not persisted —
# captures are ephemeral helpers for the current browser session.
_capture_store: dict = {}


def _cors_resp(resp):
    """Add CORS headers that allow credentialed cross-origin requests."""
    origin = request.headers.get("Origin", "")
    resp.headers["Access-Control-Allow-Origin"] = origin or "*"
    resp.headers["Access-Control-Allow-Credentials"] = "true"
    return resp


@app.route("/api/bookmarklet/capture", methods=["POST", "OPTIONS"])
def api_bookmarklet_capture():
    """Receive a product capture from the browser bookmarklet (cross-origin)."""
    if request.method == "OPTIONS":
        r = app.make_default_options_response()
        r.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return _cors_resp(r)
    if not current_user():
        return _cors_resp(jsonify(error="not logged in — open ACERT in another tab first")), 401
    data = request.get_json(silent=True) or {}
    if not data.get("url"):
        return _cors_resp(jsonify(error="url required")), 400
    _capture_store.setdefault(current_user(), []).append(data)
    return _cors_resp(jsonify(ok=True, queued=len(_capture_store[current_user()])))


@app.route("/api/captures")
def api_captures():
    """Return and clear all pending bookmarklet captures for the current user."""
    if not current_user():
        return jsonify(error="not logged in"), 401
    items = _capture_store.pop(current_user(), [])
    return jsonify(items=items)


@app.route("/bookmarklet-relay")
def bookmarklet_relay():
    """Relay page for the bookmarklet — loaded same-origin so it can POST to
    our API without hitting the product page's CSP connect-src restrictions."""
    return render_template("bookmarklet_relay.html")


@app.route("/api/orders/from_capture", methods=["POST"])
def api_order_from_capture():
    """Create a fully-populated draft order row from a bookmarklet capture."""
    if not current_user():
        return jsonify(error="not logged in"), 401
    db = get_db()
    email = current_user()
    data = request.get_json(silent=True) or {}

    # Inherit project from the user's most recent draft (same as + button)
    last = db.execute(
        "SELECT project_id FROM orders WHERE user_email=? AND status='draft'"
        " ORDER BY id DESC LIMIT 1", (email,)).fetchone()
    project_id = last["project_id"] if last else None

    # Resolve vendor: prefer existing DB entry by domain, else create from capture
    link = data.get("url", "")
    dom = domain_of(link)
    vendor_id = None
    if dom:
        all_v = fetch_vendors(db)
        matched = next((v for v in all_v if v["domain"] == dom), None)
        if matched:
            vendor_id = matched["id"]
            # Backfill placeholder name with catalog / capture data
            cat = quotes.catalog_entry_for(dom)
            want_name = (cat or {}).get("name") or data.get("vendor_name") or ""
            if want_name and matched["name"] == matched.get("domain"):
                updates = {}
                if want_name: updates["name"] = want_name
                for k in ("address", "phone"):
                    src = (cat or {}).get(k) or data.get(k, "")
                    if src and not matched.get(k):
                        updates[k] = src
                if updates:
                    set_cl = ", ".join(f"{k}=?" for k in updates)
                    db.execute(f"UPDATE vendors SET {set_cl} WHERE id=?",
                               list(updates.values()) + [vendor_id])
        elif data.get("vendor_name") or quotes.catalog_entry_for(dom):
            cat = quotes.catalog_entry_for(dom)
            vname   = (cat or {}).get("name") or data.get("vendor_name") or dom
            vaddr   = (cat or {}).get("address") or data.get("address") or ""
            vphone  = (cat or {}).get("phone")   or data.get("phone")   or ""
            vsite   = (cat or {}).get("website") or data.get("website") or dom
            cur = db.execute(
                "INSERT OR IGNORE INTO vendors"
                " (name, address, website, phone, tax_exempt_filed) VALUES (?,?,?,?,0)",
                (vname, vaddr, vsite, vphone))
            if cur.lastrowid:
                vendor_id = cur.lastrowid
            else:
                row = db.execute("SELECT id FROM vendors WHERE name=?",
                                 (vname,)).fetchone()
                vendor_id = row["id"] if row else None

    cost = _normalise_cost(data.get("price", ""))
    desc = (data.get("description") or "")[:200].strip()

    cur = db.execute(
        "INSERT INTO orders (user_email, description, link, vendor_id, project_id, cost)"
        " VALUES (?,?,?,?,?,?)",
        (email, desc, link, vendor_id, project_id, cost))
    db.commit()
    return jsonify(ok=True, order_id=cur.lastrowid)


@app.context_processor
def inject_globals():
    # Domains of quote-storage providers (Dropbox, SharePoint, OneDrive …)
    # derived from vendor_catalog.yaml so JS quoteProvider() stays in sync.
    quote_domains = [
        d
        for entry in quotes._load_catalog().get("vendors", [])
        if "quote_storage" in entry
        for d in entry.get("domains", [])
    ]
    # Bookmarklet JS — URL injected at render time so it points to this server
    try:
        bm_api   = url_for("api_bookmarklet_capture", _external=True)
        bm_relay = url_for("bookmarklet_relay",        _external=True)
    except Exception:
        bm_api = bm_relay = ""
    from urllib.parse import quote as _urlquote
    bookmarklet = "javascript:" + _urlquote(
        _BOOKMARKLET
        .replace("__CAPTURE_URL__", bm_api)
        .replace("__RELAY_URL__",   bm_relay)
    )
    return {"app_version": __version__,
            "ms_auth": _CONFIG.get("auth_provider") == "microsoft",
            "quote_storage_domains": quote_domains,
            "bookmarklet": bookmarklet}


# ── Microsoft Entra ID (Azure AD) auth ───────────────────────────────────────
def _build_msal_app():
    import msal  # only imported when MS auth is active
    return msal.ConfidentialClientApplication(
        client_id=os.environ["ORDERAPP_CLIENT_ID"],
        client_credential=os.environ["ORDERAPP_CLIENT_SECRET"],
        authority=(
            "https://login.microsoftonline.com/"
            + os.environ["ORDERAPP_TENANT_ID"]
        ),
    )


@app.route("/auth/microsoft")
def ms_login():
    try:
        msal_app = _build_msal_app()
    except KeyError as e:
        return (f"Microsoft auth is not configured: missing env var {e}. "
                "See README §Microsoft auth."), 500
    flow = msal_app.initiate_auth_code_flow(
        scopes=["openid", "profile", "email"],
        redirect_uri=url_for("ms_callback", _external=True),
    )
    session["ms_flow"] = flow
    return redirect(flow["auth_uri"])


@app.route("/auth/callback")
def ms_callback():
    flow = session.pop("ms_flow", {})
    try:
        msal_app = _build_msal_app()
    except KeyError as e:
        return render_template("login.html",
                               error=f"Server misconfiguration: missing env var {e}.")
    try:
        result = msal_app.acquire_token_by_auth_code_flow(flow, request.args)
    except ValueError as e:
        return render_template("login.html", error=f"Authentication failed: {e}")

    if "error" in result:
        return render_template("login.html",
                               error=result.get("error_description") or result["error"])

    claims = result.get("id_token_claims", {})
    email = (claims.get("preferred_username") or claims.get("email") or "").strip().lower()
    if not email or "@" not in email:
        return render_template("login.html",
                               error="Could not retrieve your email address from Microsoft.")

    db = get_db()
    ms_cfg = _CONFIG.get("microsoft", {})
    domain = email.split("@")[1]
    allowed_domains = [d.lower() for d in ms_cfg.get("allowed_domains", [])]
    is_domain_ok = domain in allowed_domains
    is_email_ok = bool(
        db.execute("SELECT 1 FROM allowed_emails WHERE email = ?", (email,)).fetchone()
    )

    if not (is_domain_ok or is_email_ok):
        return render_template(
            "login.html",
            error=f"{email} is not authorized. Ask an existing user to add you on the Users tab.")

    # Auto-add domain-authorized users so they appear in the Users tab
    if is_domain_ok and not is_email_ok:
        db.execute(
            "INSERT OR IGNORE INTO allowed_emails (email, added_by, added_at) VALUES (?,?,?)",
            (email, "microsoft-auth", now_iso()))
        log_change(db, None, "email", None, email,
                   table_name="allowed_emails", by="microsoft-auth")
        db.commit()

    session["email"] = email
    return redirect(url_for("orders"))

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

_ip_attempts = {}        # ip -> count of failed logins with disallowed emails
_IP_BLOCK_THRESHOLD = 5  # attempts before the IP is blocked

# ------------------------------------------------------------------ db

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript("""
    CREATE TABLE IF NOT EXISTS vendors (
        id INTEGER PRIMARY KEY,
        name TEXT NOT NULL UNIQUE,
        address TEXT DEFAULT '',
        website TEXT DEFAULT '',
        phone TEXT DEFAULT '',
        tax_exempt_filed INTEGER NOT NULL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY,
        name TEXT NOT NULL UNIQUE,
        notes TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY,
        user_email TEXT NOT NULL,               -- locked after submission
        description TEXT NOT NULL DEFAULT '',
        link TEXT NOT NULL DEFAULT '',
        vendor_id INTEGER REFERENCES vendors(id) ON DELETE SET NULL,
        project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
        use_note TEXT NOT NULL DEFAULT '',
        cost TEXT NOT NULL DEFAULT '',
        quantity INTEGER NOT NULL DEFAULT 1,
        status TEXT NOT NULL DEFAULT 'draft',         -- 'draft' | 'submitted'
        order_status TEXT NOT NULL DEFAULT 'submitted',-- fulfillment status
        submitted_at TEXT                             -- locked after submission
    );
    CREATE TABLE IF NOT EXISTS trackers (
        order_id INTEGER NOT NULL REFERENCES orders(id) ON DELETE CASCADE,
        email TEXT NOT NULL,
        UNIQUE (order_id, email)
    );
    CREATE TABLE IF NOT EXISTS order_history (
        id INTEGER PRIMARY KEY,
        order_id INTEGER,
        changed_by TEXT NOT NULL,
        changed_at TEXT NOT NULL,
        field TEXT NOT NULL,
        old_value TEXT,
        new_value TEXT,
        table_name TEXT NOT NULL DEFAULT 'orders'
    );
    CREATE TABLE IF NOT EXISTS allowed_emails (
        id INTEGER PRIMARY KEY,
        email TEXT NOT NULL UNIQUE,
        added_by TEXT NOT NULL DEFAULT 'system',
        added_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS blocked_ips (
        id INTEGER PRIMARY KEY,
        ip TEXT NOT NULL UNIQUE,
        blocked_at TEXT NOT NULL,
        attempts INTEGER NOT NULL DEFAULT 0,
        note TEXT NOT NULL DEFAULT ''
    );
    """)
    # Column-level migrations (idempotent — exception = already exists)
    for stmt in [
        "ALTER TABLE order_history ADD COLUMN table_name TEXT NOT NULL DEFAULT 'orders'",
        "ALTER TABLE vendors ADD COLUMN address TEXT DEFAULT ''",
        "ALTER TABLE orders ADD COLUMN cost TEXT NOT NULL DEFAULT ''",
        "ALTER TABLE orders ADD COLUMN order_status TEXT NOT NULL DEFAULT 'submitted'",
        "ALTER TABLE orders ADD COLUMN quantity INTEGER NOT NULL DEFAULT 1",
    ]:
        try:
            db.execute(stmt)
        except Exception:
            pass

    # Fix order_history if it still carries the old NOT NULL / FK constraint on
    # order_id (SQLite can't drop constraints, so rename + recreate + copy).
    old = db.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name='order_history'"
    ).fetchone()
    if old and "order_id INTEGER NOT NULL" in (old[0] or ""):
        db.executescript("""
            ALTER TABLE order_history RENAME TO _order_history_v1;
            CREATE TABLE order_history (
                id         INTEGER PRIMARY KEY,
                order_id   INTEGER,
                changed_by TEXT NOT NULL,
                changed_at TEXT NOT NULL,
                field      TEXT NOT NULL,
                old_value  TEXT,
                new_value  TEXT,
                table_name TEXT NOT NULL DEFAULT 'orders'
            );
            INSERT INTO order_history
                SELECT id, order_id, changed_by, changed_at, field,
                       old_value, new_value, COALESCE(table_name, 'orders')
                FROM _order_history_v1;
            DROP TABLE _order_history_v1;
        """)

    db.commit()
    db.close()


init_db()

# ------------------------------------------------------------------ helpers

def now_iso():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


_MERGE_WINDOW = timedelta(minutes=30)


def log_change(db, record_id, field, old, new, table_name='orders', by=None):
    """Record one field change in order_history.

    If the same user changed the same field on the same record within the
    last 30 minutes, update that existing row's new_value and changed_at
    instead of inserting a new one.  This collapses rapid edits (e.g. typing
    a description character by character) into a single A → Z entry.
    """
    who = by or current_user() or 'system'
    cutoff = (datetime.now(timezone.utc) - _MERGE_WINDOW).isoformat(timespec="seconds")

    existing = db.execute(
        "SELECT id FROM order_history"
        " WHERE order_id IS ? AND table_name=? AND field=? AND changed_by=?"
        " AND changed_at>=? ORDER BY id DESC LIMIT 1",
        (record_id or None, table_name, field, who, cutoff)
    ).fetchone()

    if existing:
        db.execute(
            "UPDATE order_history SET new_value=?, changed_at=? WHERE id=?",
            (None if new is None else str(new), now_iso(), existing["id"])
        )
    else:
        db.execute(
            "INSERT INTO order_history (order_id, changed_by, changed_at, field,"
            " old_value, new_value, table_name) VALUES (?,?,?,?,?,?,?)",
            (record_id or None, who, now_iso(), field,
             None if old is None else str(old),
             None if new is None else str(new),
             table_name)
        )


def vendor_incomplete(v):
    """A vendor is flagged if website, phone, or tax-exemption filing is missing."""
    return not ((v["website"] or "").strip() and (v["phone"] or "").strip()
                and v["tax_exempt_filed"])


def domain_of(url_or_site):
    """Bare registrable-ish domain: 'https://www.digikey.com/x' -> 'digikey.com'."""
    s = (url_or_site or "").strip()
    if not s:
        return ""
    if "://" not in s:
        s = "https://" + s
    host = (urlparse(s).hostname or "").lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def fetch_vendors(db):
    rows = db.execute("SELECT * FROM vendors ORDER BY name COLLATE NOCASE").fetchall()
    return [dict(r, incomplete=vendor_incomplete(r), domain=domain_of(r["website"]))
            for r in rows]


def fetch_projects(db):
    return db.execute("SELECT * FROM projects ORDER BY name COLLATE NOCASE").fetchall()


def trackers_for(db, order_ids):
    out = {oid: [] for oid in order_ids}
    if order_ids:
        marks = ",".join("?" * len(order_ids))
        for r in db.execute(
                f"SELECT order_id, email FROM trackers WHERE order_id IN ({marks}) ORDER BY email",
                list(order_ids)):
            out[r["order_id"]].append(r["email"])
    return out


def current_user():
    return session.get("email")


def login_required(view):
    from functools import wraps

    @wraps(view)
    def wrapped(*a, **kw):
        if not current_user():
            return redirect(url_for("login"))
        return view(*a, **kw)
    return wrapped


def order_visible_to(db, order_id, email):
    """User may touch an order if they created it or track it."""
    return db.execute(
        """SELECT o.* FROM orders o
           LEFT JOIN trackers t ON t.order_id = o.id AND t.email = ?
           WHERE o.id = ? AND (o.user_email = ? OR t.email IS NOT NULL)""",
        (email, order_id, email)).fetchone()

def get_client_ip():
    return request.headers.get("X-Forwarded-For", request.remote_addr).split(",")[0].strip()


# ------------------------------------------------------------------ auth

@app.route("/login", methods=["GET", "POST"])
def login():
    if _CONFIG.get("auth_provider") == "microsoft":
        # In Microsoft mode the GET shows the MS button; POST shouldn't occur
        # (the button is a link, not a form submit) but redirect safely if it does.
        if request.method == "POST":
            return redirect(url_for("ms_login"))
        return render_template("login.html")

    error = None
    if request.method == "POST":
        ip = get_client_ip()
        db = get_db()
        if db.execute("SELECT 1 FROM blocked_ips WHERE ip = ?", (ip,)).fetchone():
            return render_template("login.html",
                error="This IP address has been blocked after repeated failed attempts. "
                      "Contact an existing user to unblock it from the Users tab.")
        email = request.form.get("email", "").strip().lower()
        if not EMAIL_RE.match(email):
            error = "Enter a valid email address."
        elif db.execute("SELECT 1 FROM allowed_emails WHERE email = ?", (email,)).fetchone():
            session["email"] = email
            return redirect(url_for("orders"))
        else:
            count = _ip_attempts.get(ip, 0) + 1
            _ip_attempts[ip] = count
            if count >= _IP_BLOCK_THRESHOLD:
                db.execute(
                    "INSERT OR IGNORE INTO blocked_ips (ip, blocked_at, attempts) VALUES (?,?,?)",
                    (ip, now_iso(), count))
                db.execute("UPDATE blocked_ips SET attempts=?, blocked_at=? WHERE ip=?",
                           (count, now_iso(), ip))
                log_change(db, None, "ip", None, ip,
                           table_name="blocked_ips", by="system")
                db.commit()
                error = ("This IP has been blocked after too many failed attempts. "
                         "Contact an existing user to unblock it from the Users tab.")
            else:
                remaining = _IP_BLOCK_THRESHOLD - count
                error = (f"That email is not authorized. "
                         f"({remaining} attempt{'s' if remaining != 1 else ''} remaining before this IP is blocked.)")
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ------------------------------------------------------------------ pages

@app.route("/")
def index():
    return redirect(url_for("orders"))


@app.route("/debug/health")
def debug_health():
    """Diagnostic endpoint — returns JSON showing what works and what doesn't.
    Remove this route once the production issue is resolved."""
    import traceback as _tb
    result = {"version": __version__, "ok": True, "checks": {}}

    # 1. DB connection + schema
    try:
        db = get_db()
        cols = {row[1] for row in db.execute("PRAGMA table_info(orders)")}
        result["checks"]["db_orders_cols"] = sorted(cols)
    except Exception:
        result["ok"] = False
        result["checks"]["db_error"] = _tb.format_exc()

    # 2. Catalog
    try:
        cat = quotes._load_catalog()
        result["checks"]["catalog_vendors"] = len(cat.get("vendors", []))
    except Exception:
        result["ok"] = False
        result["checks"]["catalog_error"] = _tb.format_exc()

    # 3. inject_globals in real request context
    try:
        ig = inject_globals()
        result["checks"]["inject_globals_keys"] = list(ig.keys())
        result["checks"]["bookmarklet_len"] = len(ig.get("bookmarklet", ""))
    except Exception:
        result["ok"] = False
        result["checks"]["inject_globals_error"] = _tb.format_exc()

    # 4. Fetch vendors (used in orders page)
    try:
        db = get_db()
        vendors = fetch_vendors(db)
        result["checks"]["vendor_count"] = len(vendors)
    except Exception:
        result["ok"] = False
        result["checks"]["fetch_vendors_error"] = _tb.format_exc()

    # 5. Orders query + field access (simulates orders page render)
    try:
        db = get_db()
        rows = db.execute("SELECT * FROM orders LIMIT 5").fetchall()
        result["checks"]["sample_orders"] = len(rows)
        if rows:
            sample = dict(rows[0])
            result["checks"]["sample_order_keys"] = sorted(sample.keys())
            # Test fmt_cost on each
            for r in rows:
                fmt_cost(r["cost"])
    except Exception:
        result["ok"] = False
        result["checks"]["orders_query_error"] = _tb.format_exc()

    # 6. Template render test (renders orders.html with empty drafts)
    try:
        html = render_template("orders.html", tab="orders", drafts=[],
                               vendors=[], projects=[], trackers={})
        result["checks"]["template_render"] = f"ok ({len(html)} chars)"
    except Exception:
        result["ok"] = False
        result["checks"]["template_render_error"] = _tb.format_exc()

    return jsonify(result)


@app.route("/orders")
@login_required
def orders():
    db = get_db()
    email = current_user()
    drafts = db.execute(
        "SELECT * FROM orders WHERE user_email = ? AND status = 'draft' ORDER BY id",
        (email,)).fetchall()
    return render_template(
        "orders.html", tab="orders", drafts=drafts,
        vendors=fetch_vendors(db), projects=fetch_projects(db),
        trackers=trackers_for(db, [d["id"] for d in drafts]))


@app.route("/orders/new", methods=["POST"])
@login_required
def new_row():
    db = get_db()
    email = current_user()
    last = db.execute(
        "SELECT project_id FROM orders WHERE user_email = ? AND status = 'draft'"
        " ORDER BY id DESC LIMIT 1", (email,)).fetchone()
    project_id = last["project_id"] if last else None
    db.execute("INSERT INTO orders (user_email, project_id) VALUES (?,?)", (email, project_id))
    db.commit()
    return redirect(url_for("orders"))


@app.route("/orders/<int:oid>/delete", methods=["POST"])
@login_required
def delete_row(oid):
    db = get_db()
    db.execute("DELETE FROM orders WHERE id = ? AND user_email = ? AND status = 'draft'",
               (oid, current_user()))
    db.commit()
    return redirect(url_for("orders"))


@app.route("/api/orders/<int:oid>/delete", methods=["POST"])
@login_required
def api_delete_order(oid):
    """Permanently delete any order visible to the user; log all fields cleared."""
    db = get_db()
    email = current_user()
    order = order_visible_to(db, oid, email)
    if order is None:
        return jsonify(error="not found"), 404
    # Log every meaningful field as cleared so the history table shows what was lost
    for field in ("description", "link", "vendor_id", "project_id",
                  "use_note", "cost", "order_status", "status"):
        val = order[field]
        if val is not None and str(val).strip():
            log_change(db, oid, field, val, None)
    log_change(db, oid, "deleted", None, "record deleted")
    db.execute("DELETE FROM orders WHERE id = ?", (oid,))
    db.commit()
    return jsonify(ok=True)


@app.route("/orders/submit", methods=["POST"])
@login_required
def submit_orders():
    db = get_db()
    ts = now_iso()
    ids = [r["id"] for r in db.execute(
        "SELECT id FROM orders WHERE user_email = ? AND status = 'draft'",
        (current_user(),))]
    for oid in ids:
        log_change(db, oid, "status", "draft", "submitted")
    db.execute(
        "UPDATE orders SET status = 'submitted', order_status = 'submitted', submitted_at = ? "
        "WHERE user_email = ? AND status = 'draft'",
        (ts, current_user()))
    db.commit()
    return redirect(url_for("submitted"))


@app.route("/submitted")
@login_required
def submitted():
    db = get_db()
    email = current_user()
    rows = db.execute(
        """SELECT DISTINCT o.* FROM orders o
           LEFT JOIN trackers t ON t.order_id = o.id
           WHERE o.status = 'submitted' AND (o.user_email = ? OR t.email = ?)
           ORDER BY o.submitted_at DESC, o.id DESC""",
        (email, email)).fetchall()
    return render_template(
        "submitted.html", tab="submitted", rows=rows,
        vendors=fetch_vendors(db), projects=fetch_projects(db),
        trackers=trackers_for(db, [r["id"] for r in rows]))


@app.route("/vendors", methods=["GET", "POST"])
@login_required
def vendors():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            db.execute(
                "INSERT OR IGNORE INTO vendors (name, address, website, phone, tax_exempt_filed) "
                "VALUES (?,?,?,?,?)",
                (name, request.form.get("address", "").strip(),
                 request.form.get("website", "").strip(),
                 request.form.get("phone", "").strip(),
                 1 if request.form.get("tax_exempt_filed") else 0))
            db.commit()
        return redirect(url_for("vendors"))
    return render_template("vendors.html", tab="vendors", vendors=fetch_vendors(db))


@app.route("/vendors/<int:vid>/update", methods=["POST"])
@login_required
def update_vendor(vid):
    db = get_db()
    db.execute(
        "UPDATE vendors SET name=?, address=?, website=?, phone=?, tax_exempt_filed=? WHERE id=?",
        (request.form.get("name", "").strip(),
         request.form.get("address", "").strip(),
         request.form.get("website", "").strip(),
         request.form.get("phone", "").strip(),
         1 if request.form.get("tax_exempt_filed") else 0, vid))
    db.commit()
    return redirect(url_for("vendors"))


@app.route("/users", methods=["GET", "POST"])
@login_required
def users():
    db = get_db()
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if EMAIL_RE.match(email):
            cur = db.execute(
                "INSERT OR IGNORE INTO allowed_emails (email, added_by, added_at) VALUES (?,?,?)",
                (email, current_user(), now_iso()))
            if cur.rowcount:
                log_change(db, 0, "email", None, email, table_name="allowed_emails")
            db.commit()
        return redirect(url_for("users"))
    emails = db.execute("SELECT * FROM allowed_emails ORDER BY added_at DESC").fetchall()
    blocked = db.execute("SELECT * FROM blocked_ips ORDER BY blocked_at DESC").fetchall()
    return render_template("users.html", tab="users", emails=emails, blocked=blocked,
                           threshold=_IP_BLOCK_THRESHOLD)


@app.route("/users/<int:uid>/remove", methods=["POST"])
@login_required
def remove_user(uid):
    db = get_db()
    row = db.execute("SELECT email FROM allowed_emails WHERE id = ?", (uid,)).fetchone()
    if row:
        log_change(db, 0, "email", row["email"], None, table_name="allowed_emails")
        db.execute("DELETE FROM allowed_emails WHERE id = ?", (uid,))
        db.commit()
    return redirect(url_for("users"))


@app.route("/blocked-ips/<int:bid>/unblock", methods=["POST"])
@login_required
def unblock_ip(bid):
    db = get_db()
    row = db.execute("SELECT ip FROM blocked_ips WHERE id = ?", (bid,)).fetchone()
    if row:
        ip = row["ip"]
        log_change(db, 0, "ip", ip, None, table_name="blocked_ips")
        db.execute("DELETE FROM blocked_ips WHERE id = ?", (bid,))
        db.commit()
        _ip_attempts.pop(ip, None)
    return redirect(url_for("users"))


@app.route("/vendors/<int:vid>/delete", methods=["POST"])
@login_required
def delete_vendor(vid):
    db = get_db()
    v = db.execute("SELECT * FROM vendors WHERE id = ?", (vid,)).fetchone()
    if v:
        log_change(db, 0, "name", v["name"], None, table_name="vendors")
        db.execute("DELETE FROM vendors WHERE id = ?", (vid,))
        db.commit()
    return redirect(url_for("vendors"))


@app.route("/history")
@login_required
def history():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM order_history ORDER BY id DESC LIMIT 300"
    ).fetchall()
    return render_template("history.html", tab="history", rows=rows)


@app.route("/api/vendors", methods=["POST"])
@login_required
def api_create_vendor():
    """Create a vendor via JSON (from quote auto-detection). Returns {id, name, incomplete}."""
    data = request.get_json(silent=True) or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify(error="name required"), 400
    db = get_db()
    cur = db.execute(
        "INSERT OR IGNORE INTO vendors (name, address, website, phone, tax_exempt_filed) VALUES (?,?,?,?,0)",
        (name, data.get("address", "").strip(),
         data.get("website", "").strip(), data.get("phone", "").strip()))
    db.commit()
    if cur.lastrowid:
        vid = cur.lastrowid
    else:
        row = db.execute("SELECT id FROM vendors WHERE name = ?", (name,)).fetchone()
        vid = row["id"] if row else None
    v = db.execute("SELECT * FROM vendors WHERE id = ?", (vid,)).fetchone()
    vd = dict(v, incomplete=vendor_incomplete(v), domain=domain_of(v["website"]))
    return jsonify(id=vd["id"], name=vd["name"], incomplete=vd["incomplete"])


@app.route("/api/vendors/<int:vid>/patch", methods=["PATCH"])
@login_required
def api_patch_vendor(vid):
    """Update phone/website on a vendor from quote-extracted info."""
    db = get_db()
    v = db.execute("SELECT * FROM vendors WHERE id = ?", (vid,)).fetchone()
    if v is None:
        return jsonify(error="not found"), 404
    data = request.get_json(silent=True) or {}
    address = data.get("address", v["address"])
    phone   = data.get("phone",   v["phone"])
    website = data.get("website", v["website"])
    db.execute("UPDATE vendors SET address=?, phone=?, website=? WHERE id=?",
               (address, phone, website, vid))
    db.commit()
    return jsonify(ok=True)


@app.route("/projects", methods=["GET", "POST"])
@login_required
def projects():
    db = get_db()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            db.execute("INSERT OR IGNORE INTO projects (name, notes) VALUES (?,?)",
                       (name, request.form.get("notes", "").strip()))
            db.commit()
        return redirect(url_for("projects"))
    return render_template("projects.html", tab="projects",
                           projects=fetch_projects(db))


@app.route("/projects/<int:pid>/update", methods=["POST"])
@login_required
def update_project(pid):
    db = get_db()
    db.execute("UPDATE projects SET name=?, notes=? WHERE id=?",
               (request.form.get("name", "").strip(),
                request.form.get("notes", "").strip(), pid))
    db.commit()
    return redirect(url_for("projects"))

# ------------------------------------------------------------------ autosave API

# Everything is editable at any time EXCEPT who submitted (user_email) and
# when (submitted_at). Every change is written to order_history.
EDITABLE_FIELDS = {"description", "link", "vendor_id", "project_id", "use_note", "cost",
                   "quantity", "order_status"}


@app.route("/api/orders/<int:oid>", methods=["POST"])
@login_required
def api_save(oid):
    db = get_db()
    email = current_user()
    order = order_visible_to(db, oid, email)
    if order is None:
        return jsonify(error="not found"), 404
    if order["status"] == "draft" and order["user_email"] != email:
        return jsonify(error="not yours"), 403

    data = request.get_json(silent=True) or {}
    sets, vals = [], []
    for field, value in data.items():
        if field not in EDITABLE_FIELDS:
            continue
        if field in ("vendor_id", "project_id"):
            value = int(value) if str(value).strip() else None
        elif field == "cost":
            value = _normalise_cost(value)
        elif field == "quantity":
            try:
                value = max(1, int(value))
            except (ValueError, TypeError):
                value = 1
        if value != order[field]:
            log_change(db, oid, field, order[field], value)
            sets.append(f"{field} = ?")
            vals.append(value)
    if sets:
        vals.append(oid)
        db.execute(f"UPDATE orders SET {', '.join(sets)} WHERE id = ?", vals)
        db.commit()
    return jsonify(ok=True)


@app.route("/api/orders/<int:oid>/quote_vendor", methods=["POST"])
@login_required
def api_quote_vendor(oid):
    """The link points at a quote PDF on Dropbox or SharePoint/OneDrive:
    fetch it, read it, and set the vendor from the quote."""
    db = get_db()
    order = order_visible_to(db, oid, current_user())
    if order is None:
        return jsonify(error="not found"), 404

    link = (request.get_json(silent=True) or {}).get("link", "").strip()
    provider = quotes.classify_link(link)
    if provider is None:
        return jsonify(matched=False, message="Not a Dropbox/SharePoint link."), 400
    try:
        pdf_bytes = quotes.fetch_quote_pdf(link, provider)
        text = quotes.extract_text(pdf_bytes)
    except quotes.QuoteError as e:
        return jsonify(matched=False, provider=provider, message=str(e))

    # Extract price from the PDF regardless of vendor match outcome
    extracted_price = quotes.extract_net_price(text)
    if extracted_price and extracted_price != (order["cost"] or ""):
        log_change(db, oid, "cost", order["cost"], extracted_price)
        db.execute("UPDATE orders SET cost = ? WHERE id = ?", (extracted_price, oid))
        db.commit()

    all_vendors = fetch_vendors(db)
    vendor, hints = quotes.match_vendor(text, all_vendors)
    if vendor is None:
        extracted = quotes.extract_vendor_info(text)
        fuzzy = []
        if extracted and extracted.get("name"):
            fuzzy = quotes.fuzzy_match_vendors(extracted["name"], all_vendors)
        # Show popup whenever we have anything useful: fuzzy DB matches,
        # an address block from the PDF, or domain hints from the text.
        if fuzzy or extracted or hints:
            safe_fuzzy = [
                {k: v[k] for k in ("id", "name", "domain", "incomplete", "score")}
                for v in fuzzy
            ]
            return jsonify(matched=False, provider=provider,
                           fuzzy_candidates=safe_fuzzy,
                           extracted=extracted,
                           hint_domains=hints,
                           price=extracted_price)
        return jsonify(matched=False, provider=provider,
                       price=extracted_price,
                       message="Quote read, but no vendor information found in the PDF.")

    # Silently backfill any empty/placeholder fields from this quote.
    if (not vendor.get("address") or not vendor.get("phone")
            or vendor["name"] == vendor.get("domain")):
        extracted = quotes.extract_vendor_info(text)
        if extracted:
            updates = {}
            if not vendor.get("address") and extracted.get("address"):
                updates["address"] = extracted["address"]
            if not vendor.get("phone") and extracted.get("phone"):
                updates["phone"] = extracted["phone"]
            # Replace domain-as-name placeholder with the real company name.
            if vendor["name"] == vendor.get("domain") and extracted.get("name"):
                updates["name"] = extracted["name"]
            if updates:
                set_clause = ", ".join(f"{k}=?" for k in updates)
                db.execute(f"UPDATE vendors SET {set_clause} WHERE id=?",
                           list(updates.values()) + [vendor["id"]])
                db.commit()
                # Re-read so the response reflects the backfilled name.
                refreshed = db.execute(
                    "SELECT name FROM vendors WHERE id=?", (vendor["id"],)).fetchone()
                vendor_name = refreshed["name"] if refreshed else vendor["name"]
            else:
                vendor_name = vendor["name"]
    else:
        vendor_name = vendor["name"]

    if order["vendor_id"] != vendor["id"]:
        log_change(db, oid, "vendor_id", order["vendor_id"], vendor["id"])
        db.execute("UPDATE orders SET vendor_id = ? WHERE id = ?", (vendor["id"], oid))
        db.commit()
    return jsonify(matched=True, provider=provider,
                   vendor_id=vendor["id"], vendor_name=vendor_name,
                   incomplete=vendor["incomplete"],
                   price=extracted_price)


@app.route("/api/orders/<int:oid>/trackers", methods=["POST"])
@login_required
def api_add_tracker(oid):
    db = get_db()
    if order_visible_to(db, oid, current_user()) is None:
        return jsonify(error="not found"), 404
    email = (request.get_json(silent=True) or {}).get("email", "").strip().lower()
    if not EMAIL_RE.match(email):
        return jsonify(error="invalid email"), 400
    cur = db.execute("INSERT OR IGNORE INTO trackers (order_id, email) VALUES (?,?)",
                     (oid, email))
    if cur.rowcount:
        log_change(db, oid, "tracker", None, email)
    # Trackers automatically get login access
    acur = db.execute(
        "INSERT OR IGNORE INTO allowed_emails (email, added_by, added_at) VALUES (?,?,?)",
        (email, current_user(), now_iso()))
    if acur.rowcount:
        log_change(db, 0, "email", None, email, table_name="allowed_emails")
    db.commit()
    return jsonify(ok=True, email=email)


@app.route("/api/orders/<int:oid>/trackers", methods=["DELETE"])
@login_required
def api_remove_tracker(oid):
    db = get_db()
    if order_visible_to(db, oid, current_user()) is None:
        return jsonify(error="not found"), 404
    email = (request.get_json(silent=True) or {}).get("email", "").strip().lower()
    cur = db.execute("DELETE FROM trackers WHERE order_id = ? AND email = ?",
                     (oid, email))
    if cur.rowcount:
        log_change(db, oid, "tracker", email, None)
    db.commit()
    return jsonify(ok=True)


@app.route("/api/orders/<int:oid>/link_vendor", methods=["POST"])
@login_required
def api_link_vendor(oid):
    """Look up vendor by domain for a plain product URL.

    If the domain matches a known vendor, returns matched=True.
    Otherwise fetches the vendor's homepage to extract contact info and
    returns the same popup-compatible response as api_quote_vendor.
    """
    db = get_db()
    order = order_visible_to(db, oid, current_user())
    if order is None:
        return jsonify(error="not found"), 404

    data = request.get_json(silent=True) or {}
    link = data.get("link", "").strip()
    if not link:
        return jsonify(error="link required"), 400

    # Follow redirects first so shortlinks (mou.sr, amzn.to, …) resolve to
    # the real vendor domain before we look up or fetch homepage info.
    final_url = quotes.resolve_redirect(link)
    dom = domain_of(final_url) or domain_of(link)
    if not dom:
        return jsonify(error="invalid URL"), 400

    all_vendors = fetch_vendors(db)
    matched = next((v for v in all_vendors if v["domain"] == dom), None)
    if matched:
        # Backfill name/address/phone from catalog when vendor was auto-created
        # with the domain as placeholder name (e.g. "ebay.com" → "eBay").
        cat = quotes.catalog_entry_for(dom)
        if cat and matched["name"] == matched.get("domain"):
            updates = {k: cat[k] for k in ("name", "address", "phone")
                       if cat.get(k)}
            if updates:
                set_clause = ", ".join(f"{k}=?" for k in updates)
                db.execute(f"UPDATE vendors SET {set_clause} WHERE id=?",
                           list(updates.values()) + [matched["id"]])
                db.commit()
                matched = dict(matched, **updates)
        if order["vendor_id"] != matched["id"]:
            log_change(db, oid, "vendor_id", order["vendor_id"], matched["id"])
            db.execute("UPDATE orders SET vendor_id=? WHERE id=?",
                       (matched["id"], oid))
            db.commit()
        return jsonify(matched=True, vendor_id=matched["id"],
                       vendor_name=matched["name"], incomplete=matched["incomplete"])

    # Not in DB — get contact info.
    # extract_vendor_from_html checks vendor_catalog.yaml first, so known
    # vendors (Mouser, DigiKey, …) return instantly without any HTTP fetch.
    # For unknown vendors we try the homepage for JSON-LD Organisation data.
    html = None
    if not quotes.catalog_entry_for(dom):
        for homepage in (f"https://www.{dom}", f"https://{dom}"):
            html = quotes.fetch_html(homepage)
            if html:
                break
    extracted = quotes.extract_vendor_from_html(html or "", dom)

    fuzzy = []
    if extracted.get("name") and extracted["name"] != dom:
        fuzzy = quotes.fuzzy_match_vendors(extracted["name"], all_vendors)
    safe_fuzzy = [
        {k: v[k] for k in ("id", "name", "domain", "incomplete", "score")}
        for v in fuzzy
    ]
    return jsonify(matched=False, extracted=extracted,
                   fuzzy_candidates=safe_fuzzy, hint_domains=[dom])


@app.route("/api/orders/<int:oid>/fetch_price", methods=["POST"])
@login_required
def api_fetch_price(oid):
    """Scrape item price from a product page URL and save it on the order."""
    db = get_db()
    order = order_visible_to(db, oid, current_user())
    if order is None:
        return jsonify(error="not found"), 404
    # Accept link from request body (may be ahead of the autosave debounce)
    data = request.get_json(silent=True) or {}
    link = data.get("link", "").strip() or (order["link"] or "").strip()
    if not link:
        return jsonify(ok=False, message="no link on this order")
    if quotes.classify_link(link):
        return jsonify(ok=False, message="use the quote vendor button for quote links")
    # Follow shortlinks so the real product page is fetched
    link = quotes.resolve_redirect(link)
    price = quotes.fetch_item_price(link)
    if price is None:
        return jsonify(ok=False, message="could not extract price from this page")
    if price != (order["cost"] or ""):
        log_change(db, oid, "cost", order["cost"], price)
        db.execute("UPDATE orders SET cost = ? WHERE id = ?", (price, oid))
        db.commit()
    return jsonify(ok=True, price=price)


# ------------------------------------------------------------------ Excel export

def _xlsx_response(wb, filename):
    from io import BytesIO
    from flask import send_file
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _make_workbook(title, headers, rows_iter):
    """Build a styled openpyxl Workbook.

    headers: list of (label, column_width) tuples
    rows_iter: iterable of row-value tuples matching the header count
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ACCENT   = "0E6E6B"
    ALT_FILL = "E3EFEE"
    RULE     = "C9D2CF"

    thin   = Side(style="thin", color=RULE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor=ACCENT)
    a_fill = PatternFill("solid", fgColor=ALT_FILL)

    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    for col, (label, width) in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = h_fill
        c.alignment = Alignment(vertical="center")
        c.border = border
        ws.column_dimensions[c.column_letter].width = width

    for ri, values in enumerate(rows_iter, 2):
        alt = (ri % 2 == 1)
        for col, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=col, value=val)
            if alt:
                c.fill = a_fill
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = border
            # Format numeric cost columns
            if isinstance(val, float):
                c.number_format = '#,##0.00'

    return wb


def _cost_val(raw):
    """Convert raw cost string to float if possible, else keep as string."""
    if not raw:
        return ""
    try:
        return float(raw)
    except (ValueError, TypeError):
        return raw


@app.route("/export/<string:view>.xlsx")
@login_required
def export_xlsx(view):
    db  = get_db()
    email = current_user()

    if view == "submitted":
        vendor_map  = {v["id"]: v["name"] for v in fetch_vendors(db)}
        project_map = {p["id"]: p["name"] for p in fetch_projects(db)}
        rows = db.execute(
            """SELECT DISTINCT o.* FROM orders o
               LEFT JOIN trackers t ON t.order_id = o.id
               WHERE o.status = 'submitted' AND (o.user_email = ? OR t.email = ?)
               ORDER BY o.submitted_at DESC, o.id DESC""",
            (email, email)).fetchall()
        headers = [("ID", 5), ("Submitted", 12), ("By", 24), ("Description", 30),
                   ("Link", 42), ("Vendor", 20), ("Project", 18), ("Use", 22),
                   ("Cost ($)", 12), ("Order Status", 14)]
        def _rows():
            for r in rows:
                yield (r["id"], (r["submitted_at"] or "")[:10], r["user_email"],
                       r["description"], r["link"],
                       vendor_map.get(r["vendor_id"], ""),
                       project_map.get(r["project_id"], ""),
                       r["use_note"], _cost_val(r["cost"]),
                       r["order_status"] or "submitted")
        wb = _make_workbook("Submitted Orders", headers, _rows())
        return _xlsx_response(wb, "submitted_orders.xlsx")

    elif view == "drafts":
        vendor_map  = {v["id"]: v["name"] for v in fetch_vendors(db)}
        project_map = {p["id"]: p["name"] for p in fetch_projects(db)}
        rows = db.execute(
            "SELECT * FROM orders WHERE user_email = ? AND status = 'draft' ORDER BY id",
            (email,)).fetchall()
        headers = [("ID", 5), ("Description", 30), ("Link", 42), ("Vendor", 20),
                   ("Project", 18), ("Use", 22), ("Cost ($)", 12)]
        def _rows():
            for r in rows:
                yield (r["id"], r["description"], r["link"],
                       vendor_map.get(r["vendor_id"], ""),
                       project_map.get(r["project_id"], ""),
                       r["use_note"], _cost_val(r["cost"]))
        wb = _make_workbook("Draft Orders", headers, _rows())
        return _xlsx_response(wb, "draft_orders.xlsx")

    elif view == "vendors":
        vendors = fetch_vendors(db)
        headers = [("Vendor", 24), ("Address", 36), ("Website", 22),
                   ("Phone", 16), ("Tax Exempt Filed", 18)]
        def _rows():
            for v in vendors:
                yield (v["name"], v["address"], v["website"], v["phone"],
                       "Yes" if v["tax_exempt_filed"] else "No")
        wb = _make_workbook("Vendors", headers, _rows())
        return _xlsx_response(wb, "vendors.xlsx")

    elif view == "projects":
        projects = fetch_projects(db)
        headers = [("Project", 28), ("Notes", 50)]
        def _rows():
            for p in projects:
                yield (p["name"], p["notes"])
        wb = _make_workbook("Projects", headers, _rows())
        return _xlsx_response(wb, "projects.xlsx")

    else:
        return ("Unknown export view.", 404)


if __name__ == "__main__":
    app.run(debug=True)
